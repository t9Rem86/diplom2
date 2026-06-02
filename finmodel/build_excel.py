# -*- coding: utf-8 -*-
"""
Построить Excel-финмодель склада с формулами через openpyxl.

Архитектура (по образцу anthropics/financial-services xlsx-author):
  - Лист «Параметры»  — входные данные (синий шрифт) + производные (Excel-формулы)
  - Лист «Денежные потоки» — каждый месяц = одна строка, ВСЕ расчёты — формулы Excel
  - Лист «ИТОГ»       — NPV, IRR, ROI и другие KPI через Excel-формулы
  - Лист «Чувствительность» — структура для сценарного анализа

Правило: синий шрифт = введённое число, чёрный шрифт = формула.
Входные параметры меняются на листе «Параметры» → всё пересчитывается автоматически.

Использование:
    python scripts/build_excel.py --project <name>
"""
import sys
import argparse
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print('{"error": "openpyxl не установлен. Запустите: pip install openpyxl"}')
    sys.exit(1)

from finmodel.config import PROJECTS_DIR, REPORTS_DIR
from finmodel.finmodel import ProjectParams, _date_str


# ── Стили (по соглашению xlsx-author) ────────────────────────────────────────

def _font(color="000000", bold=False, size=11):
    return Font(name="Calibri", size=size, color=color, bold=bold)

F_INPUT   = _font("0000FF")             # синий  = hardcoded ввод
F_FORMULA = _font("000000")             # чёрный = формула
F_LINK    = _font("008000")             # зелёный = ссылка на другой лист
F_HDR     = _font("FFFFFF", bold=True)  # белый жирный = заголовок секции
F_SUBHDR  = _font("000000", bold=True)  # чёрный жирный = подзаголовок

FILL_HDR    = PatternFill("solid", fgColor="1F4E79")  # тёмно-синий
FILL_SUBHDR = PatternFill("solid", fgColor="D9E1F2")  # светло-синий
FILL_INPUT  = PatternFill("solid", fgColor="F2F2F2")  # светло-серый
FILL_OUTPUT = PatternFill("solid", fgColor="BDD7EE")  # средний синий
FILL_WHITE  = PatternFill("solid", fgColor="FFFFFF")
FILL_ODD    = PatternFill("solid", fgColor="F7F9FC")   # нечётные строки CF

ALIGN_R = Alignment(horizontal="right",  vertical="center")
ALIGN_L = Alignment(horizontal="left",   vertical="center")
ALIGN_C = Alignment(horizontal="center", vertical="center")

FMT_RUB = '#,##0'
FMT_PCT = '0.00%'
FMT_INT = '0'
FMT_RPCT = '#,##0.00%'


def _bd(style="thin", color="BFBFBF"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

BD_THIN   = _bd("thin")
BD_MEDIUM = _bd("medium", "4472C4")


def cell(ws, row, col, value=None, formula=None,
         font=None, fill=None, align=None, fmt=None, border=BD_THIN):
    c = ws.cell(row=row, column=col)
    c.value = formula if formula is not None else value
    if font:   c.font   = font
    if fill:   c.fill   = fill
    if align:  c.alignment = align
    if fmt:    c.number_format = fmt
    if border: c.border = border
    return c


# ── Параметры sheet ───────────────────────────────────────────────────────────
#
# Строка → (имя_поля, метка, единица, формат)
INPUT_ROWS = [
    (4,  "area_sqm",           "Площадь объекта",                    "кв.м.",             FMT_RUB),
    (5,  "construction_cost",  "Себестоимость строительства (с НДС)", "руб.",              FMT_RUB),
    (6,  "loan_share",         "Доля кредита",                        "",                  FMT_PCT),
    (7,  "rental_rate_sqm",    "Арендная ставка",                     "руб./кв.м./мес.",   FMT_RUB),
    (8,  "inflation_rate",     "Годовая инфляция (индексация аренды)","",                  FMT_PCT),
    (9,  "cbr_rate",           "Ставка ЦБ (ставка дисконтирования)",  "",                  FMT_PCT),
    (10, "loan_rate",          "Кредитная ставка",                    "",                  FMT_PCT),
    (11, "construction_months","Срок строительства",                  "мес.",              FMT_INT),
    (12, "horizon_months",     "Горизонт планирования",               "мес.",              FMT_INT),
    (13, "vat_rate",           "Ставка НДС",                          "",                  FMT_PCT),
    (14, "income_tax_rate",    "Ставка налога на прибыль",            "",                  FMT_PCT),
    (15, "property_tax_rate",  "Ставка налога на имущество",          "в год",             FMT_PCT),
    (16, "operating_cost_pct", "Операционные расходы",                "% от выручки б/НДС",FMT_PCT),
    (17, "useful_life_years",  "Срок полезного использования",        "лет",               FMT_INT),
]

# Производные: (строка, формула, метка, единица)
DERIVED_ROWS = [
    (20, "=C5/(1+C13)",               "Себестоимость без НДС",          "руб."),
    (21, "=C5-C20",                   "НДС строительства",               "руб."),
    (22, "=C5*C6",                    "Сумма кредита",                   "руб."),
    (23, "=C5*(1-C6)",                "Собственные средства",            "руб."),
    (24, "=C5/C4",                    "Стоимость 1 кв.м.",               "руб."),
    (25, "=C12-C11",                  "Период эксплуатации",             "мес."),
    (26, "=-PMT(C10/12,C12-C11,C5*C6)","Аннуитетный платёж (в мес.)",  "руб./мес."),
    (27, "=C20/(C17*12)",             "Амортизация (в мес.)",            "руб./мес."),
    (28, "=(C5-C5/(1+C13))/3",        "НДС к возмещению (1-3 мес. экспл.)", "руб./мес."),
]


def build_params(ws, params: ProjectParams):
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20

    # Строки 1-3: заголовок + название/дата
    ws.merge_cells("A1:C1")
    cell(ws, 1, 1, value="ВХОДНЫЕ ПАРАМЕТРЫ ПРОЕКТА",
         font=F_HDR, fill=FILL_HDR, align=ALIGN_L, border=BD_MEDIUM)
    ws.row_dimensions[1].height = 24

    cell(ws, 2, 1, value="Название проекта", font=F_FORMULA, fill=FILL_INPUT, align=ALIGN_L)
    cell(ws, 2, 2, value="", font=F_FORMULA, fill=FILL_INPUT)
    cell(ws, 2, 3, value=params.name, font=F_INPUT, fill=FILL_INPUT, align=ALIGN_R)

    # Дата начала — выбирается пользователем в /new-project, идёт из JSON.
    # Внутри params.start_date это "ГГГГ-ММ", отображаем как "01.ММ.ГГГГ".
    _y, _m = params.start_date.split("-")
    _start_display = f"01.{int(_m):02d}.{int(_y):04d}"
    cell(ws, 3, 1, value="Дата начала строительства", font=F_FORMULA, fill=FILL_INPUT, align=ALIGN_L)
    cell(ws, 3, 2, value="ДД.ММ.ГГГГ", font=F_FORMULA, fill=FILL_INPUT)
    cell(ws, 3, 3, value=_start_display, font=F_INPUT, fill=FILL_INPUT, align=ALIGN_R)

    # Входные параметры (строки 4-17)
    pvals = {f: getattr(params, f) for f, *_ in [(r[1],) + r[2:] for r in INPUT_ROWS]}
    pvals = vars(params)  # Use all params

    for row, field, label, unit, fmt in INPUT_ROWS:
        cell(ws, row, 1, value=label,              font=F_FORMULA, fill=FILL_INPUT, align=ALIGN_L, border=None)
        cell(ws, row, 2, value=unit,               font=F_FORMULA, fill=FILL_INPUT, align=ALIGN_L, border=None)
        cell(ws, row, 3, value=getattr(params, field),
             font=F_INPUT, fill=FILL_INPUT, align=ALIGN_R, fmt=fmt)

    # Разделитель производных параметров
    ws.merge_cells("A18:C18")
    cell(ws, 18, 1, value="РАСЧЁТНЫЕ (ПРОИЗВОДНЫЕ) ПАРАМЕТРЫ",
         font=F_HDR, fill=FILL_HDR, align=ALIGN_L, border=BD_MEDIUM)
    ws.merge_cells("A19:C19")
    cell(ws, 19, 1,
         value="Изменяйте только синие ячейки выше — формулы ниже пересчитаются автоматически.",
         font=_font("7F7F7F"), fill=FILL_WHITE, align=ALIGN_L, border=None)

    for row, formula, label, unit in DERIVED_ROWS:
        cell(ws, row, 1, value=label, font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_L, border=None)
        cell(ws, row, 2, value=unit,  font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_L, border=None)
        cell(ws, row, 3, formula=formula,
             font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_R, fmt=FMT_RUB)


# ── «Денежные потоки» sheet ───────────────────────────────────────────────────
#
# Столбцы (1-22 = A-V):
# A Month | B Date | C Phase |
# D RevGross | E RevNet | F VATout | G VATin | H VATnet |
# I Opex | J Depr | K PropTax |
# L EBIT | M Interest | N EBT | O IncTax | P NetProfit |
# Q Capex | R Drawdown | S Principal | T LoanBal |
# U InvCF | V CumCF

CF_HEADERS = [
    "Мес.", "Дата", "Фаза",
    "Выручка с НДС", "Выручка без НДС",
    "НДС выходной", "НДС входной", "НДС нетто",
    "Опер. расходы", "Амортизация", "Налог на имущ.",
    "EBIT", "Проценты", "EBT",
    "Налог на прибыль", "Чистая прибыль",
    "Капвложения", "Транш кредита", "Погашение кредита",
    "Остаток кредита", "CF инвестора", "Накопл. CF",
]

# Индексы столбцов (1-based)
cM, cD, cPH = 1, 2, 3          # Month, Date, Phase
cRG, cRN    = 4, 5              # Revenue gross/net
cVo, cVi, cVn = 6, 7, 8        # VAT out/in/net
cOp, cDp, cKt = 9, 10, 11      # Opex, Depr, PropTax
cEB, cIn, cEB2 = 12, 13, 14    # EBIT, Interest, EBT
cIT, cNP    = 15, 16            # IncTax, NetProfit
cCX, cLD, cPR = 17, 18, 19     # Capex, LoanDrawdown, Principal
cLB         = 20                # LoanBalance
cIC, cCC    = 21, 22            # InvCF, CumCF

P = "Параметры"   # имя листа параметров


def cl(col_idx):
    """Column index → Excel letter."""
    return get_column_letter(col_idx)


def build_cf(ws, params: ProjectParams):
    Tc = params.construction_months
    Th = params.horizon_months
    sy, sm = map(int, params.start_date.split("-"))

    # Ширина столбцов
    widths = [6, 10, 16, 16, 16, 14, 14, 14,
              14, 13, 14, 14, 14, 14, 16, 16,
              13, 13, 17, 16, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[cl(i)].width = w

    # Заголовок (строка 1)
    for col, hdr in enumerate(CF_HEADERS, 1):
        cell(ws, 1, col, value=hdr,
             font=F_HDR, fill=FILL_HDR, align=ALIGN_C)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "D2"

    # Строки данных (строка 2 = месяц 1)
    for t in range(1, Th + 1):
        r = t + 1
        date_s = _date_str(sy, sm, t)
        fill = FILL_ODD if t % 2 else FILL_WHITE

        # A: номер месяца
        cell(ws, r, cM, value=t, font=F_FORMULA, fill=fill, align=ALIGN_C)
        # B: дата (строковая метка, не формула)
        cell(ws, r, cD, value=date_s, font=F_FORMULA, fill=fill, align=ALIGN_C)

        if t <= Tc:
            _cf_construction(ws, r, fill)
        else:
            _cf_operations(ws, r, fill)

    # Строка ИТОГО
    tr = Th + 2
    cell(ws, tr, 1, value="ИТОГО ЗА ГОРИЗОНТ",
         font=F_SUBHDR, fill=FILL_OUTPUT, align=ALIGN_L, border=BD_MEDIUM)
    for c_idx in range(3, 23):
        col_l = cl(c_idx)
        cell(ws, tr, c_idx,
             formula=f"=SUM({col_l}2:{col_l}{Th+1})",
             font=F_SUBHDR, fill=FILL_OUTPUT, align=ALIGN_R,
             fmt=FMT_RUB, border=BD_MEDIUM)


def _ref(col_idx, row, fixed_col=True, fixed_row=True):
    """Вернуть ссылку типа $D$5 или D5."""
    c = ("$" if fixed_col else "") + cl(col_idx)
    r = ("$" if fixed_row else "") + str(row)
    return c + r


def _par(row_num):
    """Ссылка на ячейку C{row} листа Параметры."""
    return f"'{P}'!$C${row_num}"


def _cf_construction(ws, r, fill):
    """Формулы строительной фазы."""
    # Фаза
    cell(ws, r, cPH, value="Строительство", font=F_FORMULA, fill=fill, align=ALIGN_L)

    # D-L: нет выручки, нет расходов = 0
    for c_idx in [cRG, cRN, cVo, cVi, cVn, cOp, cDp, cKt, cEB, cIT, cNP]:
        cell(ws, r, c_idx, value=0, font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # T: Остаток кредита = накопленный транш = loan_amount/Tc * t
    # t = A{r}, Tc = Параметры!C11, loan_amount = Параметры!C22
    loan_bal = f"={_par(22)}/{_par(11)}*{cl(cM)}{r}"
    cell(ws, r, cLB, formula=loan_bal, font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # M: Проценты = T{r} * loan_rate/12
    interest = f"={cl(cLB)}{r}*{_par(10)}/12"
    cell(ws, r, cIn, formula=interest, font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # N: EBT = -проценты
    cell(ws, r, cEB2, formula=f"=-{cl(cIn)}{r}",
         font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # Q: Капвложения = construction_cost / Tc
    cell(ws, r, cCX, formula=f"={_par(5)}/{_par(11)}",
         font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # R: Транш кредита = loan_amount / Tc
    cell(ws, r, cLD, formula=f"={_par(22)}/{_par(11)}",
         font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # S: Погашение = 0
    cell(ws, r, cPR, value=0, font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # U: CF инвестора = -(equity_monthly + проценты)
    #    equity_monthly = equity / Tc = Параметры!C23 / Параметры!C11
    inv_cf = f"=-({_par(23)}/{_par(11)}+{cl(cIn)}{r})"
    cell(ws, r, cIC, formula=inv_cf, font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # V: Накопленный CF
    cum_cf = f"={cl(cIC)}{r}" if r == 2 else f"={cl(cCC)}{r-1}+{cl(cIC)}{r}"
    cell(ws, r, cCC, formula=cum_cf, font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)


def _cf_operations(ws, r, fill):
    """Формулы фазы эксплуатации."""
    # Фаза
    cell(ws, r, cPH, value="Эксплуатация", font=F_FORMULA, fill=fill, align=ALIGN_L)

    # D: Выручка с НДС = area * rental_rate * (1+inflation)^INT((t-Tc-1)/12)
    rev_g = (f"={_par(4)}*{_par(7)}"
             f"*(1+{_par(8)})^INT(({cl(cM)}{r}-{_par(11)}-1)/12)")
    cell(ws, r, cRG, formula=rev_g, font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # E: Выручка без НДС = D / (1 + vat_rate)
    cell(ws, r, cRN, formula=f"={cl(cRG)}{r}/(1+{_par(13)})",
         font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # F: НДС выходной = D - E
    cell(ws, r, cVo, formula=f"={cl(cRG)}{r}-{cl(cRN)}{r}",
         font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # G: НДС входной — возмещение из строительства в первые 3 мес. экспл.
    vat_in = (f"=IF(AND({cl(cM)}{r}>{_par(11)},"
              f"{cl(cM)}{r}<={_par(11)}+3),{_par(28)},0)")
    cell(ws, r, cVi, formula=vat_in, font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # H: НДС нетто = F - G (положительный = платим, отрицательный = возврат)
    cell(ws, r, cVn, formula=f"={cl(cVo)}{r}-{cl(cVi)}{r}",
         font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # I: Операционные расходы = E * operating_cost_pct
    cell(ws, r, cOp, formula=f"={cl(cRN)}{r}*{_par(16)}",
         font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # J: Амортизация (константа из Параметры!C27)
    cell(ws, r, cDp, formula=f"={_par(27)}",
         font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # K: Налог на имущество = MAX(0, cost_ex_vat - depr*(ops_t-1)) * rate / 12
    # ops_t = t - Tc, поэтому (ops_t-1) = t - Tc - 1 = A{r} - Параметры!C11 - 1
    prop_tax = (f"=MAX(0,{_par(20)}-{_par(27)}"
                f"*MAX(0,{cl(cM)}{r}-{_par(11)}-1))"
                f"*{_par(15)}/12")
    cell(ws, r, cKt, formula=prop_tax, font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # L: EBIT = E - I - J - K
    cell(ws, r, cEB, formula=f"={cl(cRN)}{r}-{cl(cOp)}{r}-{cl(cDp)}{r}-{cl(cKt)}{r}",
         font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # M: Проценты = остаток кредита прошлого месяца * ставка/12
    cell(ws, r, cIn, formula=f"={cl(cLB)}{r-1}*{_par(10)}/12",
         font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # N: EBT = EBIT - Проценты
    cell(ws, r, cEB2, formula=f"={cl(cEB)}{r}-{cl(cIn)}{r}",
         font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # O: Налог на прибыль = MAX(0, EBT * ставка)
    cell(ws, r, cIT, formula=f"=MAX(0,{cl(cEB2)}{r}*{_par(14)})",
         font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # P: Чистая прибыль = EBT - Налог
    cell(ws, r, cNP, formula=f"={cl(cEB2)}{r}-{cl(cIT)}{r}",
         font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # Q: Капвложения = 0 в фазе экспл.
    cell(ws, r, cCX, value=0, font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # R: Транш кредита = 0 в фазе экспл.
    cell(ws, r, cLD, value=0, font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # S: Погашение = MAX(0, MIN(аннуитет - проценты, остаток кредита))
    principal = f"=MAX(0,MIN({_par(26)}-{cl(cIn)}{r},{cl(cLB)}{r-1}))"
    cell(ws, r, cPR, formula=principal, font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # T: Остаток кредита = MAX(0, предыдущий - погашение)
    cell(ws, r, cLB, formula=f"=MAX(0,{cl(cLB)}{r-1}-{cl(cPR)}{r})",
         font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # U: CF инвестора = Чист.прибыль + Амортизация - Погашение - НДС_нетто
    inv_cf = f"={cl(cNP)}{r}+{cl(cDp)}{r}-{cl(cPR)}{r}-{cl(cVn)}{r}"
    cell(ws, r, cIC, formula=inv_cf, font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)

    # V: Накопленный CF
    cell(ws, r, cCC, formula=f"={cl(cCC)}{r-1}+{cl(cIC)}{r}",
         font=F_FORMULA, fill=fill, align=ALIGN_R, fmt=FMT_RUB)


# ── «ИТОГ» sheet ──────────────────────────────────────────────────────────────

def build_itog(ws, params: ProjectParams):
    Th = params.horizon_months
    last = Th + 1          # последняя строка данных в CF
    cf = "'Денежные потоки'"
    p  = f"'{P}'"

    U_col = cl(cIC)   # U — CF инвестора
    V_col = cl(cCC)   # V — накопл. CF
    D_col = cl(cRG)   # D — выручка с НДС
    E_col = cl(cRN)   # E — выручка без НДС
    I_col = cl(cOp)   # I — опер. расходы
    H_col = cl(cVn)   # H — НДС нетто
    G_col = cl(cVi)   # G — НДС входной
    O_col = cl(cIT)   # O — налог на прибыль
    K_col = cl(cKt)   # K — налог на имущество
    M_col = cl(cIn)   # M — проценты
    P_col = cl(cNP)   # P — чистая прибыль

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20

    # Заголовок
    ws.merge_cells("A1:C1")
    cell(ws, 1, 1, value="КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ ПРОЕКТА",
         font=F_HDR, fill=FILL_HDR, align=ALIGN_L, border=BD_MEDIUM)
    ws.row_dimensions[1].height = 24

    # Инфо
    for r, lbl, fml in [
        (2, "Проект",                f"={p}!$C$2"),
        (3, "Дата начала",           f"={p}!$C$3"),
        (4, "Горизонт планирования", f"={p}!$C$12&\" мес.\""),
    ]:
        cell(ws, r, 1, value=lbl, font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_L, border=None)
        cell(ws, r, 3, formula=fml, font=F_LINK, fill=FILL_WHITE, align=ALIGN_R, border=None)

    # ── KPI ──────────────────────────────────────────────────────────────────
    ws.merge_cells("A5:C5")
    cell(ws, 5, 1, value="ПОКАЗАТЕЛИ ЭФФЕКТИВНОСТИ",
         font=F_HDR, fill=FILL_HDR, align=ALIGN_L, border=BD_MEDIUM)

    kpi_rows = [
        # (row, label, formula, unit, fmt)
        (6, "NPV (чистая приведённая стоимость)",
            f"=NPV({p}!$C$9/12,{cf}!{U_col}2:{U_col}{last})",
            "руб.", FMT_RUB),
        (7, "IRR (годовая внутренняя норма доходности)",
            f"=IFERROR((1+IRR({cf}!{U_col}2:{U_col}{last}))^12-1,\"Н/Д\")",
            "%", FMT_PCT),
        (8, "ROI (рентабельность СС за горизонт)",
            f"=IFERROR(SUM({cf}!{P_col}2:{P_col}{last})/{p}!$C$23,\"Н/Д\")",
            "%", FMT_PCT),
        (9, "Срок окупаемости",
            f"=MATCH(TRUE,{cf}!{V_col}2:{V_col}{last}>=0,0)",
            "мес.", None),
        (10, "Ставка дисконтирования (ЦБ)",
             f"={p}!$C$9", "%", FMT_PCT),
    ]

    for row, lbl, fml, unit, fmt in kpi_rows:
        cell(ws, row, 1, value=lbl,  font=F_SUBHDR, fill=FILL_OUTPUT, align=ALIGN_L)
        cell(ws, row, 2, value=unit, font=F_FORMULA, fill=FILL_OUTPUT, align=ALIGN_L)
        cell(ws, row, 3, formula=fml, font=F_FORMULA, fill=FILL_OUTPUT,
             align=ALIGN_R, fmt=fmt)

    # ── Финансовые итоги ────────────────────────────────────────────────────
    ws.merge_cells("A11:C11")
    cell(ws, 11, 1, value="ФИНАНСОВЫЕ ИТОГИ ЗА ГОРИЗОНТ",
         font=F_HDR, fill=FILL_HDR, align=ALIGN_L, border=BD_MEDIUM)

    fin_rows = [
        (12, "Суммарная выручка (с НДС)",   f"=SUM({cf}!{D_col}2:{D_col}{last})", FMT_RUB),
        (13, "Суммарная выручка (без НДС)",  f"=SUM({cf}!{E_col}2:{E_col}{last})", FMT_RUB),
        (14, "Операционные расходы",          f"=SUM({cf}!{I_col}2:{I_col}{last})", FMT_RUB),
        (15, "Чистая прибыль (суммарно)",     f"=SUM({cf}!{P_col}2:{P_col}{last})", FMT_RUB),
    ]
    for row, lbl, fml, fmt in fin_rows:
        cell(ws, row, 1, value=lbl, font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_L)
        cell(ws, row, 2, value="руб.", font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_L)
        cell(ws, row, 3, formula=fml, font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_R, fmt=fmt)

    # ── Налоги ──────────────────────────────────────────────────────────────
    ws.merge_cells("A16:C16")
    cell(ws, 16, 1, value="НАЛОГИ И ПЛАТЕЖИ ЗА ГОРИЗОНТ",
         font=F_HDR, fill=FILL_HDR, align=ALIGN_L, border=BD_MEDIUM)

    tax_rows = [
        (17, "Налог на прибыль",
             f"=SUM({cf}!{O_col}2:{O_col}{last})"),
        (18, "НДС к уплате (выплаты в бюджет)",
             f"=SUMIF({cf}!{H_col}2:{H_col}{last},\">0\",{cf}!{H_col}2:{H_col}{last})"),
        (19, "Возврат НДС из строительства",
             f"=SUM({cf}!{G_col}2:{G_col}{last})"),
        (20, "Налог на имущество",
             f"=SUM({cf}!{K_col}2:{K_col}{last})"),
        (21, "Проценты по кредиту",
             f"=SUM({cf}!{M_col}2:{M_col}{last})"),
    ]
    for row, lbl, fml in tax_rows:
        cell(ws, row, 1, value=lbl, font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_L)
        cell(ws, row, 2, value="руб.", font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_L)
        cell(ws, row, 3, formula=fml, font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_R, fmt=FMT_RUB)

    # ── Финансирование ──────────────────────────────────────────────────────
    ws.merge_cells("A22:C22")
    cell(ws, 22, 1, value="СТРУКТУРА ФИНАНСИРОВАНИЯ",
         font=F_HDR, fill=FILL_HDR, align=ALIGN_L, border=BD_MEDIUM)

    fund_rows = [
        (23, "Собственные средства", f"={p}!$C$23"),
        (24, "Кредит",               f"={p}!$C$22"),
        (25, "Итого инвестиций",     f"={p}!$C$5"),
    ]
    for row, lbl, fml in fund_rows:
        cell(ws, row, 1, value=lbl, font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_L)
        cell(ws, row, 2, value="руб.", font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_L)
        cell(ws, row, 3, formula=fml, font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_R, fmt=FMT_RUB)


# ── «Чувствительность» sheet ──────────────────────────────────────────────────

def build_sensitivity(ws):
    """Сценарный анализ — заполняется автоматически при сборке отчёта.
    Все значения — формулы-ссылки на «Параметры», поэтому при изменении
    входных данных сценарии тоже пересчитываются."""
    P = "Параметры"
    I = "ИТОГ"

    # Заголовок
    ws.merge_cells("A1:E1")
    cell(ws, 1, 1, value="АНАЛИЗ ЧУВСТВИТЕЛЬНОСТИ",
         font=F_HDR, fill=FILL_HDR, align=ALIGN_L, border=BD_MEDIUM)
    ws.row_dimensions[1].height = 26

    # Ширины
    for col, w in [(1, 38), (2, 22), (3, 18), (4, 22), (5, 38)]:
        ws.column_dimensions[cl(col)].width = w

    # Заголовки колонок
    headers = ["Параметр", "Пессимист (−20%)", "Базовый",
               "Оптимист (+20%)", "Влияние на проект"]
    for col, h in enumerate(headers, 1):
        cell(ws, 3, col, value=h, font=F_SUBHDR, fill=FILL_SUBHDR, align=ALIGN_C)

    # Сценарии: (метка, формула_пес, формула_база, формула_опт, текст_влияния, формат)
    scenarios = [
        ("Себестоимость строительства, руб.",
         f"='{P}'!$C$5*1.2",  f"='{P}'!$C$5",  f"='{P}'!$C$5*0.8",
         "Рост затрат → ↓ NPV, ↑ срок окупаемости", FMT_RUB),
        ("Арендная ставка, руб./кв.м./мес.",
         f"='{P}'!$C$7*0.8",  f"='{P}'!$C$7",  f"='{P}'!$C$7*1.2",
         "Снижение ставки → ↓ выручки, ↓ NPV", FMT_RUB),
        ("Кредитная ставка",
         f"='{P}'!$C$10*1.2", f"='{P}'!$C$10", f"='{P}'!$C$10*0.8",
         "Рост ставки → ↑ процентов, ↓ чистой прибыли", FMT_PCT),
        ("Инфляция (индексация аренды)",
         f"='{P}'!$C$8*0.8",  f"='{P}'!$C$8",  f"='{P}'!$C$8*1.2",
         "Рост инфляции → ↑ выручки в будущем", FMT_PCT),
        ("Ставка ЦБ (дисконтирование)",
         f"='{P}'!$C$9*1.2",  f"='{P}'!$C$9",  f"='{P}'!$C$9*0.8",
         "Рост ставки → ↓ NPV (сильнее дисконт)", FMT_PCT),
    ]

    for r, (lbl, pes, base, opt, impact, fmt) in enumerate(scenarios, 4):
        cell(ws, r, 1, value=lbl,    font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_L)
        cell(ws, r, 2, formula=pes,  font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_R, fmt=fmt)
        cell(ws, r, 3, formula=base, font=F_SUBHDR,  fill=FILL_OUTPUT, align=ALIGN_R, fmt=fmt)
        cell(ws, r, 4, formula=opt,  font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_R, fmt=fmt)
        cell(ws, r, 5, value=impact, font=_font("595959"), fill=FILL_WHITE, align=ALIGN_L)

    # Базовые KPI (ссылки на ИТОГ)
    r = 11
    ws.merge_cells(f"A{r}:E{r}")
    cell(ws, r, 1, value="БАЗОВЫЕ KPI (ссылки на лист «ИТОГ»)",
         font=F_HDR, fill=FILL_HDR, align=ALIGN_L, border=BD_MEDIUM)
    r += 1
    for lbl, ref, unit, fmt in [
        ("NPV",                 f"='{I}'!$C$6",  "руб.", FMT_RUB),
        ("IRR (годовая)",       f"='{I}'!$C$7",  "%",    FMT_PCT),
        ("ROI",                 f"='{I}'!$C$8",  "%",    FMT_PCT),
        ("Срок окупаемости",    f"='{I}'!$C$9",  "мес.", None),
    ]:
        cell(ws, r, 1, value=lbl,     font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_L)
        cell(ws, r, 2, value=unit,    font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_L)
        cell(ws, r, 3, formula=ref,   font=F_SUBHDR,  fill=FILL_OUTPUT, align=ALIGN_R, fmt=fmt)
        r += 1

    # Инструкция
    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    cell(ws, r, 1, value="КАК ИСПОЛЬЗОВАТЬ ЭТУ ТАБЛИЦУ",
         font=F_HDR, fill=FILL_HDR, align=ALIGN_L, border=BD_MEDIUM)
    for line in [
        "1. Измените интересующий параметр на листе «Параметры» (синие ячейки C4:C17)",
        "2. Excel автоматически пересчитает базовый и сценарные значения здесь",
        "3. Откройте лист «ИТОГ» — посмотрите новые NPV, IRR, ROI",
        "4. Для полной таблицы NPV по сценариям используйте Данные → Анализ «что если» → Таблица данных",
    ]:
        r += 1
        ws.merge_cells(f"A{r}:E{r}")
        cell(ws, r, 1, value=line, font=F_FORMULA, fill=FILL_WHITE, align=ALIGN_L, border=None)



# ── Main builder ──────────────────────────────────────────────────────────────

def build(params: ProjectParams) -> Path:
    """Собрать полную Excel-финмодель и сохранить в reports/."""
    REPORTS_DIR.mkdir(exist_ok=True)
    out_path = REPORTS_DIR / f"{params.name}_finmodel.xlsx"

    wb = Workbook()

    # Удалить дефолтный лист
    default = wb.active
    wb.remove(default)

    ws_param = wb.create_sheet("Параметры")
    ws_cf    = wb.create_sheet("Денежные потоки")
    ws_itog  = wb.create_sheet("ИТОГ")
    ws_sens  = wb.create_sheet("Чувствительность")

    build_params(ws_param, params)
    build_cf(ws_cf, params)
    build_itog(ws_itog, params)
    build_sensitivity(ws_sens)

    # Авторасчёт при открытии в Excel
    wb.calculation.calcMode = "auto"

    wb.save(str(out_path))
    return out_path


def main():
    parser = argparse.ArgumentParser(description="Построить Excel-финмодель с формулами")
    parser.add_argument("--project", "-p", required=True, help="Имя проекта")
    args = parser.parse_args()

    project_file = PROJECTS_DIR / f"{args.project}.json"
    if not project_file.exists():
        import json
        print(json.dumps({"error": f"Проект не найден: {args.project}"}))
        sys.exit(1)

    params = ProjectParams.from_json(project_file)
    out_path = build(params)

    import json
    print(json.dumps({
        "status": "ok",
        "file": str(out_path),
        "sheets": ["Параметры", "Денежные потоки", "ИТОГ", "Чувствительность"],
        "horizon_months": params.horizon_months,
        "formula_rows": params.horizon_months,
        "note": "Синие ячейки на листе «Параметры» — входные данные. "
                "Измените их → Excel пересчитает модель автоматически."
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
